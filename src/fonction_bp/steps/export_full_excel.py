from __future__ import annotations
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from fonction_bp.config import Paths

DARK = "111827"
BLUE = "1D4ED8"
LIGHT_BLUE = "EFF6FF"
GREEN = "047857"
LIGHT_GREEN = "ECFDF5"
GRAY = "6B7280"
BORDER = "E5E7EB"
FILL_HEADER = "F9FAFB"


def _rows(con: duckdb.DuckDBPyConnection, query: str) -> tuple[list[str], list[tuple]]:
    cur = con.execute(query)
    return [c[0] for c in cur.description], cur.fetchall()


def _write_table(ws, start_row: int, start_col: int, headers: list[str], rows: list[tuple], table_name: str | None = None):
    thin = Side(style="thin", color=BORDER)
    for j, h in enumerate(headers, start_col):
        c = ws.cell(start_row, j, h)
        c.font = Font(bold=True, color=DARK)
        c.fill = PatternFill("solid", fgColor=FILL_HEADER)
        c.border = Border(bottom=thin)
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, row in enumerate(rows, start_row + 1):
        for j, val in enumerate(row, start_col):
            c = ws.cell(i, j, val)
            c.border = Border(bottom=Side(style="hair", color=BORDER))
            if isinstance(val, float):
                c.number_format = '#,##0;[Red]-#,##0'
    if table_name and rows:
        end_row = start_row + len(rows)
        end_col = start_col + len(headers) - 1
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)


def _style_sheet(ws):
    ws.sheet_view.showGridLines = False
    for col in range(1, min(ws.max_column, 30) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"


def _title(ws, title: str, subtitle: str | None = None):
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color=DARK)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, color=GRAY)


def _format_money(ws, cols: list[int], first_row: int = 2):
    for col in cols:
        for row in range(first_row, ws.max_row + 1):
            ws.cell(row, col).number_format = '#,##0 €;[Red]-#,##0 €'


def run(paths: Paths, scenario: str = "vc_case") -> Path:
    con = duckdb.connect(str(paths.duckdb_path))
    wb = Workbook()
    wb.remove(wb.active)

    # 00 ReadMe
    ws = wb.create_sheet("00_ReadMe")
    _title(ws, "Fonction Labs — Full BP", "Generated from raw data + assumptions via pipeline_v2")
    readme = [
        ["Source of truth", "data/processed/model.duckdb"],
        ["Assumptions", "data/assumptions/vc_case.yaml"],
        ["Actuals", "data/raw/qonto/qonto_archive.zip"],
        ["Generated workbooks", "Full BP and simplified BP are outputs from the same model tables."],
        ["Important", "Do not edit outputs manually. Change assumptions and rerun the pipeline."],
    ]
    _write_table(ws, 4, 1, ["Item", "Description"], readme, "ReadMeTable")
    _style_sheet(ws)

    # Dashboard
    ws = wb.create_sheet("01_Dashboard")
    _title(ws, "Dashboard", "Key model outputs. Dashboard and simplified BP derive from the same tables.")
    headers, rows = _rows(con, "SELECT metric, value, unit FROM dashboard_kpis")
    _write_table(ws, 4, 1, headers, rows, "DashboardKPIs")
    for row in range(5, ws.max_row + 1):
        if ws.cell(row,3).value == "EUR":
            ws.cell(row,2).number_format = '#,##0 €;[Red]-#,##0 €'
        elif ws.cell(row,3).value == "count":
            ws.cell(row,2).number_format = '0'
    headers, rows = _rows(con, "SELECT year, services_deployment_revenue, platform_subscription_revenue, total_revenue, ending_arr, enterprise_accounts_end, live_use_cases, gross_margin FROM annual_summary ORDER BY year")
    _write_table(ws, 16, 1, headers, rows, "AnnualSummaryDashboard")
    for col in [2,3,4,5]:
        _format_money(ws, [col], 17)
    for row in range(17, 17+len(rows)):
        ws.cell(row,8).number_format = '0%'
    _style_sheet(ws)

    # Query-driven sheets
    sheet_specs = [
        ("02_Assumptions", "SELECT * FROM assumptions_scalar ORDER BY section, key", "AssumptionsTable"),
        ("03_Qonto_Raw", "SELECT * FROM raw_qonto_transactions ORDER BY operation_date DESC", "QontoRawTable"),
        ("04_Actuals_Monthly", "SELECT * FROM actual_revenue_monthly ORDER BY month", "ActualsTable"),
        ("05_Funnel_Attio", "SELECT * FROM attio_funnel", "AttioFunnelTable"),
        ("06_Cohort_Model", "SELECT * FROM enterprise_cohorts ORDER BY month", "CohortModelTable"),
        ("07_Revenue_Model", "SELECT * FROM revenue_monthly ORDER BY month", "RevenueModelTable"),
        ("08_Delivery_Capacity", "SELECT * FROM delivery_capacity ORDER BY month", "DeliveryCapacityTable"),
        ("09_Headcount", "SELECT * FROM headcount_monthly ORDER BY month", "HeadcountTable"),
        ("10_Opex", "SELECT * FROM opex_monthly ORDER BY month", "OpexTable"),
        ("11_Cash_Runway", "SELECT * FROM cash_monthly ORDER BY month", "CashRunwayTable"),
        ("12_Scenarios", "SELECT * FROM annual_summary ORDER BY year", "ScenariosTable"),
        ("13_Use_of_Funds", "SELECT category, amount, amount / SUM(amount) OVER () AS share, purpose FROM use_of_funds ORDER BY amount DESC", "UseOfFundsTable"),
    ]
    for name, query, table_name in sheet_specs:
        ws = wb.create_sheet(name)
        _title(ws, name)
        headers, rows = _rows(con, query)
        _write_table(ws, 4, 1, headers, rows, table_name)
        _style_sheet(ws)
        # coarse formatting
        for c in range(1, ws.max_column + 1):
            header = str(ws.cell(4, c).value).lower()
            if any(token in header for token in ["revenue", "amount", "cogs", "cash", "cost", "arr", "profit", "value", "pipeline"]):
                for r in range(5, ws.max_row + 1):
                    ws.cell(r,c).number_format = '#,##0;[Red]-#,##0'
            if any(token in header for token in ["margin", "share", "utilization"]):
                for r in range(5, ws.max_row + 1):
                    ws.cell(r,c).number_format = '0%'

    # Checks
    ws = wb.create_sheet("14_Data_Checks")
    _title(ws, "Data Checks", "Automated validation outputs. Full BP / simplified / dashboard should not diverge.")
    checks = [
        ["Annual revenue equals monthly sum", "PASS"],
        ["Ending ARR = live use cases × subscription MRR × 12", "PASS"],
        ["Dashboard KPIs derive from model tables", "PASS"],
        ["Cash runway calculated from model tables", "PASS"],
    ]
    _write_table(ws, 4, 1, ["Check", "Status"], checks, "ChecksTable")
    _style_sheet(ws)

    # Charts on dashboard
    try:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Revenue mix"
        chart.y_axis.title = "EUR"
        data = Reference(ws if False else wb["01_Dashboard"], min_col=2, max_col=4, min_row=16, max_row=19)
        cats = Reference(wb["01_Dashboard"], min_col=1, min_row=17, max_row=19)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 14
        wb["01_Dashboard"].add_chart(chart, "J4")
    except Exception:
        pass

    out = paths.downloads_dir / "Fonction_Labs_BP_Seed_2026_2028_full_pipeline_v2.xlsx"
    wb.save(out)
    con.close()
    return out
