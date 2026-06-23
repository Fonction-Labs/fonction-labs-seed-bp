from __future__ import annotations
from pathlib import Path

import duckdb
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from fonction_bp.config import Paths

DARK = "111827"; GRAY="6B7280"; BLUE="1D4ED8"; BORDER="E5E7EB"; FILL="F9FAFB"


def _rows(con, query):
    cur=con.execute(query)
    return [c[0] for c in cur.description], cur.fetchall()


def _write(ws, r, c, headers, rows):
    thin=Side(style="thin", color=BORDER)
    for j,h in enumerate(headers,c):
        cell=ws.cell(r,j,h); cell.font=Font(bold=True,color=DARK); cell.fill=PatternFill("solid", fgColor=FILL); cell.border=Border(bottom=thin); cell.alignment=Alignment(wrap_text=True)
    for i,row in enumerate(rows,r+1):
        for j,val in enumerate(row,c):
            cell=ws.cell(i,j,val); cell.border=Border(bottom=Side(style="hair", color=BORDER))
            if isinstance(val,float): cell.number_format='#,##0;[Red]-#,##0'


def _style(ws):
    ws.sheet_view.showGridLines=False
    for col in range(1, ws.max_column+1):
        ws.column_dimensions[get_column_letter(col)].width=18
    ws.freeze_panes="A5"


def run(paths: Paths, scenario: str = "vc_case") -> Path:
    con=duckdb.connect(str(paths.duckdb_path))
    wb=Workbook(); ws=wb.active; ws.title="BP Summary"
    ws["A1"]="Fonction Labs — Simplified BP"
    ws["A1"].font=Font(size=18,bold=True,color=DARK)
    ws["A2"]="VC-facing view derived from the full model tables."
    ws["A2"].font=Font(size=10,color=GRAY)

    headers, rows = _rows(con, "SELECT metric, value, unit FROM dashboard_kpis")
    _write(ws,4,1,headers,rows)
    for i in range(5,5+len(rows)):
        if ws.cell(i,3).value == "EUR": ws.cell(i,2).number_format='#,##0 €;[Red]-#,##0 €'

    headers, rows = _rows(con, "SELECT year, services_deployment_revenue, platform_subscription_revenue, total_revenue, ending_arr, enterprise_accounts_end, live_use_cases, gross_margin FROM annual_summary ORDER BY year")
    _write(ws,17,1,headers,rows)
    for row in range(18,18+len(rows)):
        for col in [2,3,4,5]: ws.cell(row,col).number_format='#,##0 €;[Red]-#,##0 €'
        ws.cell(row,8).number_format='0%'

    headers, rows = _rows(con, "SELECT year, enterprise_accounts, live_use_cases, use_cases_per_account, ending_arr FROM year_end_milestones ORDER BY year")
    _write(ws,24,1,headers,rows)
    for row in range(25,25+len(rows)):
        ws.cell(row,4).number_format='0.0x'
        ws.cell(row,5).number_format='#,##0 €;[Red]-#,##0 €'

    headers, rows = _rows(con, "SELECT category, amount, amount / SUM(amount) OVER () AS share, purpose FROM use_of_funds ORDER BY amount DESC")
    _write(ws,31,1,headers,rows)
    for row in range(32,32+len(rows)):
        ws.cell(row,2).number_format='#,##0 €;[Red]-#,##0 €'
        ws.cell(row,3).number_format='0%'

    _style(ws)

    out=paths.downloads_dir / "Fonction_Labs_BP_Seed_2026_2028_simplified_pipeline_v2.xlsx"
    wb.save(out)
    con.close()
    return out
