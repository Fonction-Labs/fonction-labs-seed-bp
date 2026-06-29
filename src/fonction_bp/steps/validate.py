from __future__ import annotations
import json
from pathlib import Path

import duckdb
from openpyxl import load_workbook

from fonction_bp.config import Paths
from fonction_bp.utils import load_yaml, save_json


def _assert_close(name: str, actual: float, expected: float, tolerance: float, checks: list[dict]) -> None:
    actual_f = float(actual or 0)
    expected_f = float(expected or 0)
    diff = abs(actual_f - expected_f)
    ok = diff <= tolerance
    checks.append({"check": name, "actual": actual_f, "expected": expected_f, "tolerance": tolerance, "status": "PASS" if ok else "FAIL"})
    if not ok:
        raise AssertionError(f"{name}: actual={actual_f} expected={expected_f} tolerance={tolerance}")


def run(paths: Paths, scenario: str = "vc_case") -> Path:
    assumptions = load_yaml(paths.assumptions_dir / f"{scenario}.yaml")
    con = duckdb.connect(str(paths.duckdb_path))
    checks: list[dict] = []

    # Check 1: Invoiced revenue matches known total.
    invoiced_sum = con.execute("SELECT SUM(invoiced_revenue) FROM invoiced_revenue_monthly WHERE month BETWEEN DATE '2026-01-01' AND DATE '2026-06-01'").fetchone()[0]
    _assert_close("Jan-Jun 2026 invoiced revenue", invoiced_sum, float(assumptions["target_checks"]["jan_jun_2026_invoiced_revenue"]), 1.0, checks)

    # Check 2: Annual revenue equals monthly sum (internal consistency).
    bad_annual = con.execute("""
        WITH a AS (SELECT year, total_revenue FROM annual_summary),
             m AS (SELECT year, SUM(total_revenue) AS monthly_sum FROM revenue_monthly GROUP BY year)
        SELECT COUNT(*) FROM a JOIN m USING(year) WHERE ABS(a.total_revenue - m.monthly_sum) > 0.01
    """).fetchone()[0]
    checks.append({"check": "Annual revenue equals monthly sum", "actual": bad_annual, "expected": 0, "status": "PASS" if bad_annual == 0 else "FAIL"})
    if bad_annual:
        raise AssertionError("Annual revenue check failed")

    # Check 3: COGS = FDE + freelance + tokens + infra + wassym (formula check).
    bad_cogs = con.execute("""
        SELECT COUNT(*) FROM cogs_monthly
        WHERE ABS(total_cogs - (fde_cost + freelance_cost + token_cost + infra_cloud_cost + service_continuity_cogs)) > 0.01
    """).fetchone()[0]
    checks.append({"check": "COGS formula internal consistency", "actual": bad_cogs, "expected": 0, "status": "PASS" if bad_cogs == 0 else "FAIL"})
    if bad_cogs:
        raise AssertionError("COGS formula check failed")

    # Check 4: FDE headcount is non-negative and monotonically reasonable.
    neg_fde = con.execute("SELECT COUNT(*) FROM fde_headcount_plan WHERE fde_headcount < 0").fetchone()[0]
    checks.append({"check": "FDE headcount non-negative", "actual": neg_fde, "expected": 0, "status": "PASS" if neg_fde == 0 else "FAIL"})
    if neg_fde:
        raise AssertionError("Negative FDE headcount found")

    # Check 5: Cash never goes below -500k (reasonable burn check).
    min_cash = con.execute("SELECT MIN(ending_cash) FROM cash_monthly").fetchone()[0]
    checks.append({"check": "Cash runway floor (> -500k)", "actual": float(min_cash or 0), "expected": 0, "tolerance": 500000, "status": "PASS" if (min_cash or 0) > -500000 else "WARN"})

    # Check 6: Dashboard data alignment.
    data = json.loads(paths.model_outputs_json.read_text(encoding="utf-8"))
    dash_arr = next((k["value"] for k in data["kpis"] if k["metric"] == "Ending ARR Dec-2027"), None)
    if dash_arr is None:
        dash_arr = 0
    dec_2027_arr = con.execute("SELECT ending_arr FROM revenue_monthly WHERE month = DATE '2027-12-01'").fetchone()[0]
    _assert_close("Dashboard Dec-2027 ARR alignment", dash_arr, dec_2027_arr, 0.01, checks)

    # Check 7: Workbook exists and has expected sheets.
    full_path = paths.downloads_dir / "Fonction_Labs_BP_Seed_2026_2028_full_pipeline_v2.xlsx"
    simple_path = paths.downloads_dir / "Fonction_Labs_BP_Seed_2026_2028_simplified_pipeline_v2.xlsx"
    if not full_path.exists() or not simple_path.exists():
        raise FileNotFoundError("Excel outputs missing")
    wb = load_workbook(full_path, read_only=True, data_only=False)
    expected_sheets = ['00_ReadMe','01_Dashboard','02_Assumptions','03_Actuals_Summary','04_Actuals_Monthly','05_Funnel_Attio','06_Cohort_Model','07_Revenue_Model','08_Delivery_Capacity','09_Headcount','10_Opex','11_Cash_Runway','12_Scenarios','13_Use_of_Funds','14_Data_Checks']
    missing = [s for s in expected_sheets if s not in wb.sheetnames]
    checks.append({"check": "Full workbook sheets", "actual": len(missing), "expected": 0, "status": "PASS" if not missing else "FAIL"})
    if missing:
        raise AssertionError(f"Missing sheets: {missing}")

    report = {"status": "PASS", "checks": checks}
    out = paths.processed_dir / "validation_report.json"
    save_json(out, report)
    con.close()
    return out
